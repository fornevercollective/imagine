#!/usr/bin/env python3
"""
Point of Truth (PoT) generator.
Reads pot/pot.json (single source of truth) and emits all requested formats:
  md, js, json, py, txt, csv, xlsx, binary(.pkl), html

Run from repo root:
  python pot/generators/build_all.py

Or from pot/:
  python generators/build_all.py

After editing pot.json, re-run to update exports/.
"""
from __future__ import annotations
import json
import pickle
import csv
import sys
from pathlib import Path
from datetime import datetime, timezone
from html import escape as h

ROOT = Path(__file__).resolve().parents[2]
POT_JSON = ROOT / "pot" / "pot.json"
EXPORTS = ROOT / "pot" / "exports"
EXPORTS.mkdir(parents=True, exist_ok=True)

def load_pot():
    with open(POT_JSON) as f:
        return json.load(f)

def write_text(path: Path, content: str):
    path.write_text(content, encoding="utf-8")
    print(f"wrote {path.relative_to(ROOT)}")

def write_json(path: Path, data):
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"wrote {path.relative_to(ROOT)}")

def write_binary(path: Path, obj):
    with open(path, "wb") as f:
        pickle.dump(obj, f, protocol=pickle.HIGHEST_PROTOCOL)
    print(f"wrote {path.relative_to(ROOT)}")

# ---------- Generators ----------

def gen_md(pot):
    lines = []
    meta = pot["meta"]
    lines.append(f"# {meta['title']}")
    lines.append("")
    lines.append(f"**Source**: {meta['source_share_url']}")
    lines.append(f"**Version**: {meta['version']} | **Updated**: {meta['last_updated']}")
    lines.append("")
    lines.append(meta["description"])
    lines.append("")
    lines.append("## Timeline (Film Genres ↔ Literary Parallels)")
    lines.append("")
    lines.append("| Era | Year | Film Genre | Literary Parallel | Key Film | Key Tags |")
    lines.append("|-----|------|------------|-------------------|----------|----------|")
    for t in pot["timeline"]:
        tags = ", ".join(t.get("tags", [])[:5])
        lines.append(f"| {t['era']} | {t['year']} | {t['film_genre']} | {t['literary_parallel']} | {t['key_film']} | {tags} |")
    lines.append("")
    lines.append("See [loop.md](../loop.md) for the full psychological + provenance model.")
    lines.append("")
    lines.append("## Provenance Stages (Real Life → EXIF Shoot → AI)")
    lines.append("")
    for s in pot["provenance_stages"]:
        lines.append(f"### Stage {s['stage']}: {s['name']}")
        lines.append(f"Short: `{s['short']}` | EXIF: {s['exif_applicable']}")
        lines.append("")
        lines.append("Psych: " + ", ".join(s.get("psych_keywords", [])))
        lines.append("")
        if s.get("tags"):
            lines.append("Tags: " + ", ".join(f"`{t}`" for t in s["tags"]))
        lines.append("")
    lines.append("## Tag Categories (Every film/video/gif/stream/content tag)")
    lines.append("")
    for cat, tags in pot["tag_categories"].items():
        lines.append(f"### {cat}")
        lines.append(" ".join(f"`{t}`" for t in tags))
        lines.append("")
    lines.append("## Flat Tag List (for grepping / scripts)")
    lines.append("")
    lines.append(" ".join(f"`{t}`" for t in pot["all_tags_flat"]))
    lines.append("")
    lines.append("## Project Sync Snapshot")
    lines.append("")
    lines.append("Film emulation + cinematic + pinterest groups are embedded from style_presets/groups.json.")
    lines.append("")
    lines.append("## Example Source Shoot (EXIF Ground Truth)")
    lines.append("")
    lines.append("See `data/example-shoot.json` — full Leica M11 + 50/1.4 + dawn location + lighting notes + custom pot: XMP fields for the Roman Empire jaw study.")
    lines.append("")
    lines.append("## Generated Formats")
    lines.append("")
    lines.append("This .md + pot.js + pot.json + pot.py + pot.txt + pot.csv + pot.xlsx + pot.pkl + pot.html are all emitted from pot.json by generators/build_all.py")
    lines.append("")
    lines.append(f"_Generated at {datetime.now(timezone.utc).isoformat()}Z_")
    write_text(EXPORTS / "pot.md", "\n".join(lines))

def gen_js(pot):
    content = f"""// pot.js — ES module export of the Point of Truth
// Generated from pot/pot.json — do not edit directly.

export const POT = {json.dumps(pot, indent=2, ensure_ascii=False)};

export const META = POT.meta;
export const TIMELINE = POT.timeline;
export const PROVENANCE_STAGES = POT.provenance_stages;
export const TAG_CATEGORIES = POT.tag_categories;
export const ALL_TAGS = POT.all_tags_flat;

export function getTagsByCategory(cat) {{
  return TAG_CATEGORIES[cat] || [];
}}

export function getTimelineEntry(id) {{
  return TIMELINE.find(t => t.id === id);
}}

export function searchTags(q) {{
  const qq = q.toLowerCase();
  return ALL_TAGS.filter(t => t.toLowerCase().includes(qq));
}}

export default POT;
"""
    write_text(EXPORTS / "pot.js", content)

def gen_json(pot):
    # clean copy (already pretty)
    write_json(EXPORTS / "pot.json", pot)

def gen_yaml(pot):
    try:
        import yaml
    except ImportError:
        print("pyyaml not found — installing with --break-system-packages")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--break-system-packages", "pyyaml"])
        import yaml
    # Use safe_dump with block style for readability
    y = yaml.safe_dump(pot, sort_keys=False, allow_unicode=True, default_flow_style=False, width=100)
    path = EXPORTS / "pot.yaml"
    path.write_text(y, encoding="utf-8")
    print(f"wrote {path.relative_to(ROOT)}")

def gen_py(pot):
    # Emit a tiny loader that reads the sibling pot.json (keeps data in sync, valid Python always).
    # This is the pragmatic single-source way: pot.json is truth, pot.py is convenient import wrapper.
    py = '''#!/usr/bin/env python3
"""
pot.py — Python module form of the Point of Truth
Import: from pot.exports.pot import POT, TIMELINE, search_tags, etc.
Or: import pot.exports.pot as pot

The actual data lives in sibling pot.json (single source of truth).
This wrapper makes `import pot` ergonomics nice while staying 100% faithful.
"""
from __future__ import annotations
import json
from pathlib import Path

_HERE = Path(__file__).parent
with open(_HERE / "pot.json", encoding="utf-8") as f:
    POT = json.load(f)

META = POT["meta"]
TIMELINE = POT["timeline"]
PROVENANCE_STAGES = POT["provenance_stages"]
TAG_CATEGORIES = POT["tag_categories"]
ALL_TAGS = POT["all_tags_flat"]

def get_tags_by_category(cat: str):
    return TAG_CATEGORIES.get(cat, [])

def get_timeline_entry(tid: str):
    for t in TIMELINE:
        if t.get("id") == tid:
            return t
    return None

def search_tags(q: str):
    qq = q.lower()
    return [t for t in ALL_TAGS if qq in t.lower()]

def get_provenance_stage(n: int):
    for s in PROVENANCE_STAGES:
        if s.get("stage") == n:
            return s
    return None

if __name__ == "__main__":
    print("PoT version:", META["version"])
    print("Timeline entries:", len(TIMELINE))
    print("Total flat tags:", len(ALL_TAGS))
    print("Example search 'roman':", search_tags("roman")[:3])
'''
    write_text(EXPORTS / "pot.py", py)

def gen_txt(pot):
    # "raw Apache-style format"
    lines = []
    lines.append("# Film Genres: Literary Parallels Timeline — Point of Truth")
    lines.append(f"# Source: {pot['meta']['source_share_url']}")
    lines.append(f"# Version: {pot['meta']['version']}")
    lines.append("# Every film/video/gif/stream of content tag, extracted & canonicalized.")
    lines.append("# This is the Apache-style / plain text dump (one tag per line under sections).")
    lines.append("# Generated from pot.json — edit source, re-run generator.")
    lines.append("")
    lines.append("# === TIMELINE GENRE TAGS (unique across eras) ===")
    seen = set()
    for t in pot["timeline"]:
        for tag in t.get("tags", []):
            if tag not in seen:
                seen.add(tag)
                lines.append(tag)
    lines.append("")
    lines.append("# === TAG CATEGORIES ===")
    for cat, tags in pot["tag_categories"].items():
        lines.append(f"# -- {cat} --")
        for tag in tags:
            lines.append(tag)
        lines.append("")
    lines.append("# === FLAT MASTER LIST (all unique) ===")
    for tag in pot["all_tags_flat"]:
        lines.append(tag)
    lines.append("")
    lines.append("# === PROVENANCE STAGE SHORT TAGS ===")
    for s in pot["provenance_stages"]:
        lines.append(s["short"])
        for t in s.get("tags", []):
            lines.append(t)
    lines.append("")
    lines.append("# End of pot.txt (Apache-style raw tag export)")
    write_text(EXPORTS / "pot.txt", "\n".join(lines))

def gen_csv(pot):
    # multiple logical tables in one CSV? Or one big. Better: several CSVs or one with sections.
    # We'll do one master CSV with type discriminator for simplicity + easy import.
    rows = []
    # timeline rows
    for t in pot["timeline"]:
        rows.append({
            "type": "timeline",
            "id": t.get("id"),
            "era": t.get("era"),
            "year": t.get("year"),
            "film_genre": t.get("film_genre"),
            "literary_parallel": t.get("literary_parallel"),
            "key_film": t.get("key_film"),
            "tags": "|".join(t.get("tags", [])),
        })
    # provenance
    for s in pot["provenance_stages"]:
        rows.append({
            "type": "provenance_stage",
            "stage": s["stage"],
            "name": s["name"],
            "short": s["short"],
            "exif": s["exif_applicable"],
            "tags": "|".join(s.get("tags", [])),
        })
    # flat tags
    for tag in pot["all_tags_flat"]:
        rows.append({"type": "flat_tag", "tag": tag})
    # category tags
    for cat, tags in pot["tag_categories"].items():
        for tag in tags:
            rows.append({"type": "category_tag", "category": cat, "tag": tag})

    path = EXPORTS / "pot.csv"
    if rows:
        fieldnames = sorted({k for r in rows for k in r.keys()})
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)
    print(f"wrote {path.relative_to(ROOT)}")

def gen_xlsx(pot):
    # Use openpyxl; auto-install if missing (break-system-packages for dev machine)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not found — installing with --break-system-packages")
        import subprocess
        subprocess.check_call([
            sys.executable, "-m", "pip", "install",
            "--break-system-packages", "openpyxl"
        ])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Sheet 1: Timeline
    ws = wb.active
    ws.title = "Timeline"
    headers = ["id", "era", "year", "film_genre", "literary_parallel", "key_film", "tags"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(bold=True, color="FFFFFF")
    for r, t in enumerate(pot["timeline"], 2):
        ws.cell(row=r, column=1, value=t.get("id"))
        ws.cell(row=r, column=2, value=t.get("era"))
        ws.cell(row=r, column=3, value=t.get("year"))
        ws.cell(row=r, column=4, value=t.get("film_genre"))
        ws.cell(row=r, column=5, value=t.get("literary_parallel"))
        ws.cell(row=r, column=6, value=t.get("key_film"))
        ws.cell(row=r, column=7, value=", ".join(t.get("tags", [])))
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 28 if col < 7 else 60

    # Sheet 2: Provenance Loop
    ws2 = wb.create_sheet("Provenance Loop")
    headers2 = ["stage", "name", "short", "exif", "psych_keywords", "tags"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="70AD47")
    for r, s in enumerate(pot["provenance_stages"], 2):
        ws2.cell(row=r, column=1, value=s["stage"])
        ws2.cell(row=r, column=2, value=s["name"])
        ws2.cell(row=r, column=3, value=s["short"])
        ws2.cell(row=r, column=4, value=s["exif_applicable"])
        ws2.cell(row=r, column=5, value="; ".join(s.get("psych_keywords", [])))
        ws2.cell(row=r, column=6, value=", ".join(s.get("tags", [])))
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 22 if col != 5 else 45

    # Sheet 3: Tags by Category
    ws3 = wb.create_sheet("Tags by Category")
    ws3.cell(row=1, column=1, value="category").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="tag").font = Font(bold=True)
    ws3.cell(row=1, column=1).fill = PatternFill("solid", fgColor="ED7D31")
    ws3.cell(row=1, column=2).fill = PatternFill("solid", fgColor="ED7D31")
    row = 2
    for cat, tags in pot["tag_categories"].items():
        for tag in tags:
            ws3.cell(row=row, column=1, value=cat)
            ws3.cell(row=row, column=2, value=tag)
            row += 1
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 35

    # Sheet 4: Flat Tags
    ws4 = wb.create_sheet("Flat Tags")
    ws4.cell(row=1, column=1, value="tag").font = Font(bold=True)
    ws4.cell(row=1, column=1).fill = PatternFill("solid", fgColor="7030A0")
    for i, tag in enumerate(pot["all_tags_flat"], 2):
        ws4.cell(row=i, column=1, value=tag)
    ws4.column_dimensions["A"].width = 35

    # Sheet 5: Meta + Loop Summary
    ws5 = wb.create_sheet("Meta + Loop")
    ws5["A1"] = "Point of Truth — Film Genres Literary Parallels + Full Provenance Loop"
    ws5["A1"].font = Font(bold=True, size=14)
    ws5["A3"] = "Source Share"
    ws5["B3"] = pot["meta"]["source_share_url"]
    ws5["A4"] = "Version"
    ws5["B4"] = pot["meta"]["version"]
    ws5["A6"] = "See loop.md for the complete psychological model of real-life EXIF shoot → AI concept."
    ws5.column_dimensions["A"].width = 20
    ws5.column_dimensions["B"].width = 80

    # Sheet 6: Project Sync
    ws6 = wb.create_sheet("Project Sync")
    ws6["A1"] = "Synced from style_presets/groups.json + featured_templates + assets"
    r = 3
    for grp, items in pot["project_sync"]["style_presets_groups"].items():
        ws6.cell(row=r, column=1, value=grp).font = Font(bold=True)
        r += 1
        for item in items:
            ws6.cell(row=r, column=2, value=item)
            r += 1
        r += 1

    path = EXPORTS / "pot.xlsx"
    wb.save(path)
    print(f"wrote {path.relative_to(ROOT)}")

def gen_pkl(pot):
    write_binary(EXPORTS / "pot.pkl", pot)

def gen_html(pot):
    # Self-contained interactive single-file explorer
    meta = pot["meta"]
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>PoT — {meta['title']}</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
:root {{ --bg:#0f1115; --fg:#e6e6e6; --accent:#7aa2f7; --card:#1a1d24; }}
body {{ font-family: system-ui, -apple-system, Segoe UI, Roboto, sans-serif; margin:0; background:var(--bg); color:var(--fg); line-height:1.5; }}
header {{ background:#161921; padding:1rem 1.25rem; border-bottom:1px solid #333; position:sticky; top:0; z-index:10; }}
h1 {{ margin:0; font-size:1.4rem; }}
h2 {{ margin-top:2rem; border-bottom:1px solid #333; padding-bottom:.25rem; }}
.container {{ max-width:1100px; margin:0 auto; padding:1rem; }}
.card {{ background:var(--card); border:1px solid #2a2f3a; border-radius:8px; padding:1rem; margin-bottom:1rem; }}
table {{ width:100%; border-collapse:collapse; font-size:0.9rem; }}
th,td {{ padding:6px 8px; border:1px solid #333; text-align:left; vertical-align:top; }}
th {{ background:#222; position:sticky; top:0; }}
.tag {{ display:inline-block; background:#222; color:#9cc; padding:1px 6px; border-radius:3px; margin:1px; font-size:0.8rem; font-family:ui-monospace, monospace; }}
.filter {{ width:100%; padding:8px; background:#111; color:#ddd; border:1px solid #444; border-radius:4px; font-size:1rem; }}
.nav {{ display:flex; gap:8px; flex-wrap:wrap; margin:1rem 0; }}
button {{ background:#222; color:#ddd; border:1px solid #444; padding:6px 10px; border-radius:4px; cursor:pointer; }}
button:hover {{ background:#2a2f3a; }}
#loop {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(140px,1fr)); gap:8px; }}
.stage {{ background:#161921; border:1px solid #333; padding:8px; border-radius:6px; font-size:0.85rem; }}
.stage .num {{ font-weight:700; color:#7aa2f7; }}
#tagcloud {{ display:flex; flex-wrap:wrap; gap:4px; }}
.small {{ font-size:0.8rem; opacity:0.7; }}
footer {{ text-align:center; padding:2rem 1rem; opacity:0.6; font-size:0.8rem; }}
</style>
</head>
<body>
<header>
  <div class="container">
    <h1>PoT — Point of Truth</h1>
    <div class="small">{meta['source_share_url']}</div>
    <div class="small">v{meta['version']} • {meta['last_updated']}</div>
  </div>
</header>

<div class="container">

<section>
<h2>Psychological Loop (Real Life EXIF → AI Concept)</h2>
<p class="small">The full chain. Tags at every stage make the transformations explicit and traceable. See loop.md for deep dive.</p>
<div id="loop">
"""
    for s in pot["provenance_stages"]:
        html += f"""<div class="stage"><span class="num">S{s['stage']}</span> <strong>{h(s['name'])}</strong><br><span class="small">{h(s['short'])}</span><br><span class="small">EXIF: {s['exif_applicable']}</span></div>"""
    html += """
</div>
</section>

<section>
<h2>Timeline — Film Genres : Literary Parallels</h2>
<input id="timelineFilter" class="filter" placeholder="Filter timeline (era, genre, lit, film, tags)..." oninput="filterTable('timelineTable', this.value)">
<table id="timelineTable">
<thead><tr><th>Era</th><th>Year</th><th>Film Genre</th><th>Literary Parallel</th><th>Key Film</th><th>Tags</th></tr></thead>
<tbody>
"""
    for t in pot["timeline"]:
        tags_html = " ".join(f'<span class="tag">{h(tag)}</span>' for tag in t.get("tags", []))
        html += f"<tr><td>{h(t['era'])}</td><td>{t['year']}</td><td>{h(t['film_genre'])}</td><td>{h(t['literary_parallel'])}</td><td>{h(t['key_film'])}</td><td>{tags_html}</td></tr>\n"
    html += """
</tbody></table>
</section>

<section>
<h2>All Tags — Searchable / Filterable</h2>
<input id="tagFilter" class="filter" placeholder="Type to filter tags (e.g. roman, noir, kodak, exif, ghosted)..." oninput="filterTags(this.value)">
<div id="tagcloud">
"""
    for tag in pot["all_tags_flat"]:
        html += f'<span class="tag" data-tag="{h(tag)}">{h(tag)}</span>'
    html += """
</div>
<p class="small">These are every film/video/gif/stream of content tag surfaced from the source conversation + project styles + provenance model.</p>
</section>

<section>
<h2>Provenance Stages Detail</h2>
"""
    for s in pot["provenance_stages"]:
        html += f"""<div class="card"><strong>Stage {s['stage']}: {h(s['name'])}</strong> <span class="tag">{h(s['short'])}</span><br>
        <span class="small">EXIF applicable: {s['exif_applicable']}</span><br>
        Psych: {h(', '.join(s.get('psych_keywords',[])))}<br>
        Tags: {" ".join(f'<span class="tag">{h(t)}</span>' for t in s.get('tags',[])) }
        </div>"""
    html += """
</section>

<section>
<h2>Downloads — All Formats (client-side data: URLs)</h2>
<div class="nav">
<button onclick="downloadJSON()">Download pot.json</button>
<button onclick="downloadJS()">Download pot.js</button>
<button onclick="downloadTXT()">Download pot.txt (Apache-style)</button>
<button onclick="downloadCSV()">Download pot.csv</button>
<button onclick="downloadPY()">Download pot.py</button>
<button onclick="downloadPKL()">Download pot.pkl (binary)</button>
<button onclick="downloadMD()">Download pot.md</button>
<button onclick="downloadXLSX()">Download pot.xlsx (note: this page cannot produce real xlsx; use the generator)</button>
</div>
<p class="small">The .xlsx and full fidelity binary are best produced by running the Python generator (it has openpyxl + pickle). The buttons above give you the in-memory data in the other formats.</p>
</section>

<section>
<h2>Meta</h2>
<div class="card">
<pre style="white-space:pre-wrap; font-size:0.8rem; background:#111; padding:8px; border-radius:4px;">{h(json.dumps(pot['meta'], indent=2))}</pre>
</div>
</section>

</div>
<footer>
Generated from pot.json • Interactive single-file PoT explorer • Iterate by editing source + re-running generator
</footer>

<script>
// Embedded full data for client-side exports
const POT = {json.dumps(pot, ensure_ascii=False)};

function filterTable(tableId, q) {{
  const rows = document.querySelectorAll(`#${{tableId}} tbody tr`);
  const qq = q.toLowerCase();
  rows.forEach(r => {{
    const txt = r.textContent.toLowerCase();
    r.style.display = txt.includes(qq) ? '' : 'none';
  }});
}}

function filterTags(q) {{
  const tags = document.querySelectorAll('#tagcloud .tag');
  const qq = q.toLowerCase();
  tags.forEach(el => {{
    const t = el.dataset.tag.toLowerCase();
    el.style.display = t.includes(qq) ? 'inline-block' : 'none';
  }});
}}

function downloadBlob(filename, content, type) {{
  const blob = new Blob([content], {{type}});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = filename;
  a.click();
  URL.revokeObjectURL(a.href);
}}

function downloadJSON() {{ downloadBlob('pot.json', JSON.stringify(POT, null, 2), 'application/json'); }}
function downloadJS() {{
  const js = `// pot.js — generated from in-browser PoT\\nexport const POT = ${{JSON.stringify(POT, null, 2)}};\\nexport default POT;`;
  downloadBlob('pot.js', js, 'text/javascript');
}}
function downloadTXT() {{
  let txt = '# Apache-style raw tags from PoT (client export)\\n';
  txt += POT.all_tags_flat.join('\\n');
  downloadBlob('pot.txt', txt, 'text/plain');
}}
function downloadCSV() {{
  let csv = 'type,value\\n';
  POT.all_tags_flat.forEach(t => csv += `flat_tag,"${{t}}"\\n`);
  downloadBlob('pot.csv', csv, 'text/csv');
}}
function downloadPY() {{
  const py = `POT = ${{JSON.stringify(POT, null, 2)}}\\nprint('PoT loaded, version', POT.meta.version)`;
  downloadBlob('pot.py', py, 'text/x-python');
}}
function downloadPKL() {{
  alert('Binary pickle cannot be produced reliably in browser. Run the Python generator instead (exports/pot.pkl).');
}}
function downloadMD() {{
  const md = `# PoT (client export)\\n\\nSource: ${{POT.meta.source_share_url}}\\n\\nSee the full generator output for rich tables.`;
  downloadBlob('pot.md', md, 'text/markdown');
}}
function downloadXLSX() {{
  alert('Real .xlsx requires the server-side generator (openpyxl). The CSV/JSON from here can be imported into Excel.');
}}

// Keyboard: / focuses filter
document.addEventListener('keydown', e => {{
  if (e.key === '/' && document.activeElement.tagName === 'BODY') {{
    e.preventDefault();
    const f = document.getElementById('tagFilter') || document.getElementById('timelineFilter');
    if (f) f.focus();
  }}
}});
console.log('%c[PoT] Interactive explorer ready. Data embedded from pot.json', 'color:#555');
</script>
</body>
</html>
"""
    write_text(EXPORTS / "pot.html", html)

# ---------- Main ----------

def main():
    print("Loading PoT from", POT_JSON)
    pot = load_pot()
    print("Loaded version", pot["meta"]["version"], "with", len(pot["timeline"]), "timeline entries,", len(pot["all_tags_flat"]), "flat tags")

    gen_md(pot)
    gen_js(pot)
    gen_json(pot)
    gen_yaml(pot)
    gen_py(pot)
    gen_txt(pot)
    gen_csv(pot)
    gen_xlsx(pot)
    gen_pkl(pot)
    gen_html(pot)

    print("\nAll exports written to", EXPORTS)
    print("Done.")

if __name__ == "__main__":
    main()
